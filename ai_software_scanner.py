#!/usr/bin/env python3
"""
AI Software Scanner

Reads an Excel file of approved software and uses AI to determine
which ones contain embedded AI features that need security review.

Usage:
    python ai_software_scanner.py software_inventory.xlsx  # Uses MASTER Spreadsheet
    python ai_software_scanner.py software_inventory.xlsx --sheet "Dental School"  # Specific sheet
    python ai_software_scanner.py software_inventory.xlsx --all  # All sheets

Output:
    - Console summary
    - ai_scan_results.csv with all results
"""

import csv
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openai import OpenAI


def load_software_list(
    filepath: str, sheet_name: str = None, all_sheets: bool = False
) -> list[dict]:
    """Load software entries from Excel file."""
    software = []
    # wb = openpyxl.load_workbook(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if all_sheets:
        sheets_to_process = wb.sheetnames
    else:
        sheets_to_process = [sheet_name or "MASTER Spreadsheet"]

    for sheet in sheets_to_process:
        if sheet not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet}' not found, skipping...")
            continue

        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find column indices from header row
        header = [str(c).lower().strip() if c else "" for c in rows[0]]
        vendor_col = product_col = desc_col = status_col = None

        for i, col in enumerate(header):
            if "vendor" in col and "name" in col:
                vendor_col = i
            elif col == "product name":
                product_col = i
            elif col == "description":
                desc_col = i
            elif col == "status":
                status_col = i

        if vendor_col is None or product_col is None:
            continue

        # Process data rows
        for row in rows[1:]:
            if len(row) <= max(vendor_col, product_col):
                continue

            vendor = str(row[vendor_col] or "").strip()
            product = str(row[product_col] or "").strip()
            desc = (
                str(row[desc_col] or "").strip()
                if desc_col and len(row) > desc_col
                else ""
            )
            status = (
                str(row[status_col] or "").strip().upper()
                if status_col and len(row) > status_col
                else ""
            )

            if (
                not vendor
                or not product
                or vendor.lower() == "nan"
                or product.lower() == "nan"
            ):
                continue
            if status == "INACTIVE":
                continue
            if desc.lower() == "nan":
                desc = ""

            software.append(
                {
                    "vendor": vendor,
                    "product": product,
                    "description": desc,
                    "sheet": sheet,
                }
            )

    return software


def check_for_ai(client: OpenAI, entry: dict) -> dict:
    """Use OpenAI to determine if software contains AI features."""
    software_info = f"{entry['vendor']} {entry['product']}"
    if entry["description"]:
        software_info += f"\nDescription: {entry['description']}"

    prompt = f"""You are a software analyst checking if applications contain AI/ML features.

For the following software, determine if it contains any embedded AI, machine learning, 
or features that might send data to AI cloud services.

SOFTWARE:
{software_info}

Consider things like:
- Voice transcription or speech-to-text
- AI-powered assistants or chatbots (e.g., Copilot, Assistant features)
- Machine learning models for predictions/recommendations
- AI image/document processing or OCR with AI
- Natural language processing features
- Cloud-based AI APIs the software might use
- Smart/intelligent automation features
- Computer vision or image recognition

Respond in this exact format:
HAS_AI: YES or NO or UNKNOWN
CONFIDENCE: HIGH, MEDIUM, or LOW
REASON: One sentence explaining your assessment

Be conservative - if there's a reasonable chance it has AI features, say YES.
If you don't recognize the software, say UNKNOWN."""

    try:
        response = client.chat.completions.create(
            model="gpt-5.2",
            # max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.choices[0].message.content

        # DEBUG:
        # print(f"\n--- Response for {software_name} ---")
        # print(repr(text))
        # print("--- End ---")

        # Parse response
        has_ai, confidence, reason = "UNKNOWN", "LOW", "Could not determine"
        for line in text.split("\n"):
            line = line.strip()
            if line.startswith("HAS_AI:"):
                value = line.replace("HAS_AI:", "").strip().upper()
                has_ai = (
                    "YES" if "YES" in value else ("NO" if "NO" in value else "UNKNOWN")
                )
            elif line.startswith("CONFIDENCE:"):
                confidence = line.replace("CONFIDENCE:", "").strip().upper()
            elif line.startswith("REASON:"):
                reason = line.replace("REASON:", "").strip()

        return {"has_ai": has_ai, "confidence": confidence, "reason": reason}
    except Exception as e:
        return {"has_ai": "ERROR", "confidence": "N/A", "reason": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python ai_software_scanner.py <file.xlsx> [--sheet NAME | --all]")
        sys.exit(1)

    input_file = sys.argv[1]
    sheet_name = None
    all_sheets = "--all" in sys.argv

    if "--sheet" in sys.argv:
        idx = sys.argv.index("--sheet")
        if idx + 1 < len(sys.argv):
            sheet_name = sys.argv[idx + 1]

    if not Path(input_file).exists():
        print(f"Error: File '{input_file}' not found")
        sys.exit(1)

    # Load software list
    print(f"Loading from {input_file}...")
    software_list = load_software_list(input_file, sheet_name, all_sheets)
    print(f"Found {len(software_list)} software entries\n")

    if not software_list:
        print("No software found.")
        sys.exit(1)

    client = OpenAI()
    results, flagged = [], []

    for i, entry in enumerate(software_list, 1):
        name = f"{entry['vendor']} - {entry['product']}"
        print(f"[{i}/{len(software_list)}] {name}...", end=" ", flush=True)

        result = check_for_ai(client, entry)
        result.update(entry)
        results.append(result)

        if result["has_ai"] in ("YES", "UNKNOWN"):
            flagged.append(result)
            print(f"⚠️ FLAGGED ({result['has_ai']})")
        else:
            print("✓ OK")

    # Write results to CSV
    output_file = "ai_scan_results.csv"
    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "Sheet",
                "Vendor",
                "Product",
                "Description",
                "Has AI",
                "Confidence",
                "Reason",
                "Needs Review",
            ]
        )
        for r in results:
            needs_review = "YES" if r["has_ai"] in ("YES", "UNKNOWN") else "NO"
            writer.writerow(
                [
                    r["sheet"],
                    r["vendor"],
                    r["product"],
                    r["description"],
                    r["has_ai"],
                    r["confidence"],
                    r["reason"],
                    needs_review,
                ]
            )

    # Print summary
    print("\n" + "=" * 60)
    print(f"SCAN COMPLETE - {len(results)} checked, {len(flagged)} flagged")
    print(f"Results saved to: {output_file}")
    print(f"Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
