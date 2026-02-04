"""
AI Scanner Core Module

Shared functionality for scanning software for AI features using OpenAI or Azure OpenAI.
"""

import csv
from typing import Union

import openpyxl
from openai import OpenAI, AzureOpenAI


def load_software_list(
    filepath: str, sheet_name: str = None, all_sheets: bool = False
) -> list[dict]:
    """Load software entries from Excel file."""
    software = []
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if all_sheets:
        sheets_to_process = wb.sheetnames
    else:
        sheets_to_process = [sheet_name or "MASTER Spreadsheet"]

    for sheet in sheets_to_process:
        if sheet not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet}' not found, skipping...")
            continue

        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find column indices from header row
        header = [str(c).lower().strip() if c else "" for c in rows[0]]
        vendor_col = product_col = desc_col = status_col = None

        for i, col in enumerate(header):
            if "vendor" in col and "name" in col:
                vendor_col = i
            elif col == "product name":
                product_col = i
            elif col == "description":
                desc_col = i
            elif col == "status":
                status_col = i

        if vendor_col is None or product_col is None:
            continue

        # Process data rows
        for row in rows[1:]:
            if len(row) <= max(vendor_col, product_col):
                continue

            vendor = str(row[vendor_col] or "").strip()
            product = str(row[product_col] or "").strip()
            desc = (
                str(row[desc_col] or "").strip()
                if desc_col and len(row) > desc_col
                else ""
            )
            status = (
                str(row[status_col] or "").strip().upper()
                if status_col and len(row) > status_col
                else ""
            )

            if (
                not vendor
                or not product
                or "nan" in (vendor.lower(), product.lower())
            ):
                continue
            if status == "INACTIVE":
                continue
            if desc.lower() == "nan":
                desc = ""

            software.append(
                {
                    "vendor": vendor,
                    "product": product,
                    "description": desc,
                    "sheet": sheet,
                }
            )

    return software


def check_for_ai(
    client: Union[OpenAI, AzureOpenAI], model: str, entry: dict
) -> dict:
    """Use OpenAI/Azure OpenAI to determine if software contains AI features."""
    software_info = f"{entry['vendor']} {entry['product']}"
    if entry["description"]:
        software_info += f"\nDescription: {entry['description']}"

    prompt = f"""You are a software analyst checking if applications contain AI/ML features.

For the following software, determine if it contains any embedded AI, machine learning,
or features that might send data to AI cloud services.

SOFTWARE:
{software_info}

Consider things like:
- Voice transcription or speech-to-text
- AI-powered assistants or chatbots (e.g., Copilot, Assistant features)
- Machine learning models for predictions/recommendations
- AI image/document processing or OCR with AI
- Natural language processing features
- Cloud-based AI APIs the software might use
- Smart/intelligent automation features
- Computer vision or image recognition

Respond in this exact format:
HAS_AI: YES or NO or UNKNOWN
CONFIDENCE: HIGH, MEDIUM, or LOW
REASON: One concise sentence (max 255 characters) explaining your assessment

Be conservative - if there's a reasonable chance it has AI features, say YES.
If you don't recognize the software, say UNKNOWN."""

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.choices[0].message.content

        # Parse response
        has_ai, confidence, reason = "UNKNOWN", "LOW", "Could not determine"
        for line in text.split("\n"):
            line = line.strip()
            if line.startswith("HAS_AI:"):
                value = line.replace("HAS_AI:", "").strip().upper()
                has_ai = (
                    "YES" if "YES" in value else ("NO" if "NO" in value else "UNKNOWN")
                )
            elif line.startswith("CONFIDENCE:"):
                confidence = line.replace("CONFIDENCE:", "").strip().upper()
            elif line.startswith("REASON:"):
                reason = line.replace("REASON:", "").strip()

        # Truncate reason to 255 characters if needed
        if len(reason) > 255:
            reason = reason[:252].rsplit(" ", 1)[0] + "..."

        return {"has_ai": has_ai, "confidence": confidence, "reason": reason}
    except Exception as e:
        return {"has_ai": "ERROR", "confidence": "N/A", "reason": str(e)}


def scan_software(
    client: Union[OpenAI, AzureOpenAI], model: str, software_list: list[dict]
) -> tuple[list[dict], list[dict]]:
    """Scan software list for AI features and return results."""
    results, flagged = [], []

    for i, entry in enumerate(software_list, 1):
        name = f"{entry['vendor']} - {entry['product']}"
        print(f"[{i}/{len(software_list)}] {name}...", end=" ", flush=True)

        result = check_for_ai(client, model, entry)
        result.update(entry)
        results.append(result)

        if result["has_ai"] in ("YES", "UNKNOWN"):
            flagged.append(result)
            print(f"⚠️ FLAGGED ({result['has_ai']})")
        else:
            print("✓ OK")

    return results, flagged


def save_results(results: list[dict], output_file: str = "ai_scan_results.csv"):
    """Save scan results to CSV file."""
    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "Sheet",
                "Vendor",
                "Product",
                "Description",
                "Has AI",
                "Confidence",
                "Reason",
                "Needs Review",
            ]
        )
        for r in results:
            needs_review = "YES" if r["has_ai"] in ("YES", "UNKNOWN") else "NO"
            writer.writerow(
                [
                    r["sheet"],
                    r["vendor"],
                    r["product"],
                    r["description"],
                    r["has_ai"],
                    r["confidence"],
                    r["reason"],
                    needs_review,
                ]
            )
